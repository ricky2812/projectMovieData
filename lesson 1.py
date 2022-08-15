from selenium import webdriver
import time
import urllib.request
import chromedriver_autoinstaller




import pandas as pd #importing pandas module
import openpyxl as xl
from openpyxl.chart import BarChart, Reference
  


#workbook = pd.read_excel(r"C:\Users\Aadil\Desktop\Alien Brains Python workshop\Intenship\indian movies.csv", engine='xlsxreader', usecols = 'B:H')
#workbook.head()






wb = xl.load_workbook(r"C:\Users\Aadil\Desktop\Alien Brains Python workshop\Intenship\indian movies.xlsx")
sheet = wb['indian movies']


m_id = (sheet.cell(31,1)).value
m_name = (sheet.cell(31,2)).value


# now opening imdb 




#chromedriver_autoinstaller.install()

driver = webdriver.Chrome("chromedriver.exe")   #step 1, opening chrome

driver.get('https://www.imdb.com/title/'+ m_id)


# driver.find_element_by_xpath('//*[@id="suggestion-search"]').send_keys(m_name)

# driver.find_element_by_xpath('//*[@id="iconContext-magnify"]').click()


# driver.find_element_by_xpath('/html/body/div[3]/div/div[2]/div[3]/div[1]/div/div[2]/table/tbody/tr[1]/td[2]/a').click()  #clicking first result

try:
    rating = driver.find_element_by_xpath('//*[@class="sc-7ab21ed2-1 jGRxWM"]').text
    rating = float(rating)
except:
    rating = 'not defined'
    
year = driver.find_element_by_xpath('//*[@class="sc-8c396aa2-2 itZqyK"]').text
if year!='':
    year = float(year)

director = driver.find_element_by_xpath('//*[@class="ipc-metadata-list-item__list-content-item ipc-metadata-list-item__list-content-item--link"]').text
                
try:    
    plot = driver.find_element_by_xpath('//*[@class="sc-16ede01-0 fMPjMP"]').text
except:
    plot = 'Not Available'
    pass

l = driver.find_elements_by_xpath('//*[@class="ipc-inline-list ipc-inline-list--show-dividers baseAlt"]')
genres = []

for i in l:
    genres.append(i.text)

l = driver.find_elements_by_xpath('//*[@class="ipc-metadata-list-item__list-content-item ipc-metadata-list-item__list-content-item--link"]')
   
actors = []
for i in l:
    actors.append(i.text)



new_list = [m_id,m_name, rating, year, director, plot, genres, actors]
df = pd.DataFrame(new_list)
df=df.transpose()
writer = pd.ExcelWriter('details_of_movies_from_imdb.xlsx', engine='xlsxwriter')
df.to_excel(writer, sheet_name='movies', index=False)
writer.save()





